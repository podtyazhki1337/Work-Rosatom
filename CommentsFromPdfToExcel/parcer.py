import PyPDF2
import pandas as pd
from datetime import datetime
import os
import shutil
import sqlite3
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
import re


def extract_comments_from_pdf(pdf_path):
    try:
        with open(pdf_path, 'rb') as file:
            pdf_reader = PyPDF2.PdfReader(file)
            comments = []
            has_annotations = False

            print(f"\nАнализ файла: {os.path.basename(pdf_path)}")
            print(f"Количество страниц: {len(pdf_reader.pages)}")

            for page_num in range(len(pdf_reader.pages)):
                page = pdf_reader.pages[page_num]

                if '/Annots' in page:
                    has_annotations = True
                    annotations = page['/Annots']
                    print(f"Страница {page_num + 1}: найдено {len(annotations)} аннотаций")

                    for annot in annotations:
                        annotation = annot.get_object()
                        subtype = annotation.get('/Subtype', 'Unknown')

                        if subtype != '/Widget':
                            content = annotation.get('/Contents', '')
                            match = re.match(r'#([CQR])#\s*(.*)', content)
                            if match:
                                comment_type, clean_content = match.groups()
                                category = {'C': 'Contents', 'Q': 'Question', 'R': 'Requirements'}.get(comment_type,
                                                                                                       'Unknown')
                            else:
                                clean_content = content
                                category = 'Unknown'

                            comment = {
                                'Document': os.path.basename(pdf_path),
                                'Non-conformance ID': '',
                                'Chapter': '',
                                'Page': page_num + 1,
                                'Owner’s remark': clean_content,
                                'Name (Owner)': annotation.get('/T', 'Unknown'),
                                'Category of remark': category
                            }
                            comments.append(comment)
                        else:
                            print(f"Пропущена цифровая подпись (/Widget) на странице {page_num + 1}")

                else:
                    print(f"Страница {page_num + 1}: аннотации отсутствуют")

            if not has_annotations:
                print("В файле нет аннотаций")
            elif not comments:
                print("Комментариев (кроме цифровых подписей) не найдено")

            return comments

    except Exception as e:
        print(f"Ошибка при обработке файла {pdf_path}: {str(e)}")
        return []


def save_to_sql(all_comments, output_folder):
    db_path = os.path.join(output_folder, 'comments.db')
    conn = sqlite3.connect(db_path)
    cursor = conn.cursor()

    cursor.execute('''
        CREATE TABLE IF NOT EXISTS comments (
            "Document" TEXT,
            "Non-conformance ID" TEXT,
            "Chapter" TEXT,
            "Page" INTEGER,
            "Owner’s remark" TEXT,
            "Name (Owner)" TEXT,
            "Category of remark" TEXT
        )
    ''')

    for comment in all_comments:
        cursor.execute('''
            INSERT INTO comments ("Document", "Non-conformance ID", "Chapter", "Page", "Owner’s remark", "Name (Owner)", "Category of remark")
            VALUES (?, ?, ?, ?, ?, ?, ?)
        ''', (comment['Document'], comment['Non-conformance ID'], comment['Chapter'],
              comment['Page'], comment['Owner’s remark'], comment['Name (Owner)'], comment['Category of remark']))

    conn.commit()
    conn.close()
    print(f"Данные сохранены в SQL базу: {db_path}")


def save_to_excel(all_comments, template_path, output_folder):
    output_filename = f'comments_output_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    output_path = os.path.join(output_folder, output_filename)
    shutil.copy(template_path, output_path)

    # Определяем полный список из 17 столбцов, где Category of remark на 8-м месте (H)
    columns = [
        'Document', 'Non-conformance ID', 'Chapter', 'Page',
        'Owner’s remark', 'Name (Owner)', 'Empty1', 'Category of remark',
        'Empty2', 'Empty3', 'Empty4', 'Empty5', 'Empty6', 'Empty7', 'Empty8', 'Empty9', 'Empty10'
    ]

    # Создаем DataFrame с 17 столбцами
    df = pd.DataFrame(columns=columns)
    if all_comments:
        temp_df = pd.DataFrame(all_comments, columns=['Document', 'Non-conformance ID', 'Chapter',
                                                      'Page', 'Owner’s remark', 'Name (Owner)', 'Category of remark'])
        df = pd.concat([df, temp_df], ignore_index=True)
        df.fillna('', inplace=True)  # Заполняем пустые столбцы

    book = load_workbook(output_path)
    sheet_name = list(book.sheetnames)[0]
    sheet = book[sheet_name]

    # Проверяем объединённые ячейки и разъединяем их ниже 3-й строки
    merged_ranges = list(sheet.merged_cells.ranges)
    for merged_range in merged_ranges:
        if merged_range.min_row > 3:
            sheet.unmerge_cells(str(merged_range))
            print(f"Разъединён диапазон: {merged_range}")

    # Находим первую пустую строку после 3-й в столбце A
    start_row = 4
    while sheet[f'A{start_row}'].value is not None:
        start_row += 1

    # Записываем данные с помощью pandas
    with pd.ExcelWriter(output_path, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
        df.to_excel(writer, sheet_name=sheet_name, index=False, startrow=start_row - 1, header=False)

    # Повторно загружаем книгу для применения форматирования
    book = load_workbook(output_path)
    sheet = book[sheet_name]

    # Применяем форматирование из строки 4 к новым строкам (A-Q, 1-17)
    for row in range(start_row, start_row + len(all_comments)):
        for col in range(1, 18):  # Столбцы A-Q (1-17)
            source_cell = sheet[f"{get_column_letter(col)}4"]
            target_cell = sheet[f"{get_column_letter(col)}{row}"]
            if source_cell.has_style:
                target_cell._style = source_cell._style

    book.save(output_path)
    print(f"\nДанные сохранены в {output_path}, лист: {sheet_name}, начиная с строки {start_row}")


def process_folder(folder_path, template_path):
    all_comments = []

    if not os.path.exists(folder_path):
        print("Указанная папка не существует!")
        return

    pdf_files = [f for f in os.listdir(folder_path) if f.lower().endswith('.pdf')]

    if not pdf_files:
        print("В папке нет PDF файлов!")
        return

    for pdf_file in pdf_files:
        pdf_path = os.path.join(folder_path, pdf_file)
        comments = extract_comments_from_pdf(pdf_path)
        all_comments.extend(comments)

    if all_comments:
        save_to_excel(all_comments, template_path, folder_path)
        save_to_sql(all_comments, folder_path)
    else:
        print("Комментариев для записи не найдено")


def main():
    print("Программа для извлечения комментариев из PDF файлов")
    print("------------------------------------------------")

    script_dir = os.path.dirname(os.path.abspath(__file__))
    template_path = os.path.join(script_dir, 'template.xlsx')

    if not os.path.exists(template_path):
        print(f"Шаблон Excel не найден по пути: {template_path}")
        return

    folder_path = input("Введите путь к папке с PDF файлами: ")
    process_folder(folder_path, template_path)


if __name__ == "__main__":
    main()